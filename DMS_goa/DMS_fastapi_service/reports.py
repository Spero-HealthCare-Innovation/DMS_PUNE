from fastapi import APIRouter, Query
from typing import Optional
from DMS_goa.hive_connector import hive_connecter_execution
from io import BytesIO
import pandas as pd
from admin_web.models import *
from fastapi.responses import StreamingResponse
from datetime import datetime, timedelta
import pytz
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font 

india_tz = pytz.timezone('Asia/Kolkata')


router = APIRouter()

@router.get('/incident_report_incident_daywise')
def incident_report_incident_daywise(
    from_date : Optional[str] = Query(..., description="Start date in YYYY-MM-DD format"),
    to_date : Optional[str] = Query(..., description="End date in YYYY-MM-DD format")):
    try:
        to_date_obj = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)
        to_date_plus_one = to_date_obj.strftime("%Y-%m-%d")

        incident_data = DMS_Incident.objects.filter(inc_added_date__range = (from_date,to_date_plus_one), inc_type=1)
        data = [{"incident_id": str(incident.incident_id),"incident_datetime": incident.inc_datetime,"disaster_name": incident.disaster_type.disaster_name if incident.disaster_type else None,"incident_type": "Emergency","alert_type": ("High" if incident.alert_type == 1 else "Medium" if incident.alert_type == 2 else "Low" if incident.alert_type == 3 else "Very Low" if incident.alert_type == 4 else "" ),"responder":  ', '.join(list({DMS_Responder.objects.get(responder_id=int(k)).responder_name for notif in DMS_Notify.objects.filter(incident_id=incident, not_is_deleted=False) for k in notif.alert_type_id }))} for incident in incident_data]
        return data
    except Exception as e:
        return {"Error":"Error","msg":str(e)}



@router.get('/download_incident_report_incident_daywise')
def download_incident_report_incident_daywise(
    from_date: Optional[str] = Query(..., description="Start date in YYYY-MM-DD format"),
    to_date: Optional[str] = Query(..., description="End date in YYYY-MM-DD format")
):
    try:
        to_date_obj = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)
        to_date_plus_one = to_date_obj.strftime("%Y-%m-%d")

        incident_data = DMS_Incident.objects.filter(inc_added_date__range = (from_date,to_date_plus_one), inc_type=1)
        wb = Workbook()
        ws = wb.active
        ws.title = "Incident Report"
        headers = ["S. No.", "Incident Id", "Incident Datetime", "Disaster Name", "Incident Type", "Alert Type", "Responder","Caller Number", "Caller Name", "Call Duration", "Incident Address", "Incident District","Incident Tahsil", "Ward", "Incident Summary", "Remark"]
        ws.append(headers)
        bold_font = Font(bold=True)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = bold_font
        row_num = 2
        serial_no = 1
        for i in incident_data:
            responders = list({
                DMS_Responder.objects.get(responder_id=int(k)).responder_name
                for j in DMS_Notify.objects.filter(incident_id=i, not_is_deleted=False)
                for k in j.alert_type_id
            })
            remarks = DMS_Comments.objects.filter(incident_id=i, comm_is_deleted=False)
            remark_list = [z.comments for z in remarks] or [""]
            start_row = row_num
            for idx, remark in enumerate(remark_list):
                ws.append([serial_no if idx == 0 else "",str(i.incident_id),i.inc_datetime,i.disaster_type.disaster_name if i.disaster_type else None,"Emergency" if i.inc_type == 1 else "Non-Emergency","High" if i.alert_type == 1 else "Medium" if i.alert_type == 2 else "Low" if i.alert_type == 3 else "Very Low" if i.alert_type == 4 else None,', '.join(responders),i.caller_id.caller_no if i.caller_id else None,i.caller_id.caller_name if i.caller_id else None,i.time,i.location,i.district.dis_name if i.district else None,i.tahsil.tah_name if i.tahsil else None,i.ward.ward_name if i.ward else None,i.summary.summary if i.summary else None,remark])
                row_num += 1
            end_row = row_num - 1
            if end_row > start_row:
                for col_index in range(1, len(headers)): 
                    cell_range = f"{get_column_letter(col_index)}{start_row}:{get_column_letter(col_index)}{end_row}"
                    ws.merge_cells(cell_range)
                ws.merge_cells(start_row=start_row, end_row=end_row, start_column=1, end_column=1)
            serial_no += 1
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"incident_report_{from_date}_to_{to_date}.xlsx"
        headers = {'Content-Disposition': f'attachment; filename="{filename}"'}
        return StreamingResponse(output,media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',headers=headers)
    except Exception as e:
        return {"Error": "Error", "msg": str(e)}





@router.get("/incident_closure_report_daywise/")
def incident_report_daywise(
    from_date: Optional[str] = Query(..., description="Start date in YYYY-MM-DD format"),
    to_date: Optional[str] = Query(..., description="End date in YYYY-MM-DD format")):
    try:
        to_date_obj = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)
        to_date_plus_one = to_date_obj.strftime("%Y-%m-%d")

        closure_data = DMS_incident_closure.objects.filter(closure_added_date__range=(from_date, to_date_plus_one))
        dt = []
        for i in closure_data:
            nn={
                "incident_id": str(i.incident_id.incident_id if i.incident_id else None),
                "disaster_type": i.incident_id.disaster_type.disaster_name if i.incident_id and i.incident_id.disaster_type else None,
                "closure_acknowledge": i.closure_acknowledge,
                "closure_start_base_location": i.closure_start_base_location,
                "closure_at_scene": i.closure_at_scene,
                "closure_from_scene": i.closure_from_scene,
                "closure_back_to_base": i.closure_back_to_base,
                "closure_remark": i.closure_remark
            }
            dt.append(nn)
        return dt
    except Exception as e:
        return {"status": "error","message": str(e)}





@router.get("/download_incident_closure_report_daywise/")
def incident_report_daywise(
    from_date: Optional[str] = Query(..., description="Start date in YYYY-MM-DD format"),
    to_date: Optional[str] = Query(..., description="End date in YYYY-MM-DD format")
):
    try:
        to_date_obj = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)
        to_date_plus_one = to_date_obj.strftime("%Y-%m-%d")
        closure_data = DMS_incident_closure.objects.filter(closure_added_date__range=(from_date, to_date_plus_one))
        dt = []
        for i in closure_data:
            if i.incident_id:
                nn={
                    "incident_id": str(i.incident_id.incident_id if i.incident_id else None),
                    "Alert Source":"System Alert" if  i.incident_id.mode == 2 else "Manual Calls",
                    "disaster_type": i.incident_id.disaster_type.disaster_name if i.incident_id and i.incident_id.disaster_type else None,
                    "Alert Type": "High" if i.incident_id.alert_type == 1 else "Medium" if i.incident_id.alert_type == 2 else "Low" if i.incident_id.alert_type == 3 else "Very Low" if i.incident_id.alert_type == 4 else "Unknown",
                    "Alert Time" : i.incident_id.alert_id.alert_datetime if i.incident_id and i.incident_id.alert_id else None,
                    "Incident Dispatch Time" : i.incident_id.inc_added_date,
                    "Closure Time": i.closure_added_date,
                    "closure_acknowledge": i.closure_acknowledge,
                    "closure_start_base_location": i.closure_start_base_location,
                    "closure_at_scene": i.closure_at_scene,
                    "closure_from_scene": i.closure_from_scene,
                    "closure_back_to_base": i.closure_back_to_base,
                    "closure_remark": i.closure_remark,
                    "Caller Number" : i.incident_id.caller_id.caller_name if i.incident_id and i.incident_id.caller_id else None,
                    "Caller Name" : i.incident_id.caller_id.caller_no if i.incident_id and i.incident_id.caller_id else None,
                    "Address" : i.incident_id.location if i.incident_id else None,
                    "Lattitude" : i.incident_id.latitude if i.incident_id else None,
                    "Longitude" : i.incident_id.longitude if i.incident_id else None
                }

            
            dt.append(nn)

        df = pd.DataFrame(dt)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Incident Report')
        output.seek(0)

        filename = f"incident_report_{from_date}_to_{to_date}.xlsx"
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"'
        }

        return StreamingResponse(output, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)

    except Exception as e:
        return {"status": "error", "message": str(e)}
